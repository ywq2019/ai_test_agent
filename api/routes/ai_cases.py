"""
AI 文档驱动用例生成路由
  - /ai-cases/*  含 CRUD、diff-check、incremental-update、optimize、coverage
"""
from fastapi import APIRouter, HTTPException, Depends, BackgroundTasks, Request
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from typing import List, Dict, Any, Optional
from pathlib import Path
from api.limiter import limiter
from api.auth import get_current_user, owner_filter, check_owner, workspace_filter, workspace_filter_members, check_workspace_member, check_access
from loguru import logger

from tools.database import get_db, AICaseFile, User
from api.websocket_manager import ws_manager
from api.schemas import (
    AICaseGenerateRequest, AICaseFileResponse,
    DiffCheckRequest, IncrementalUpdateRequest,
)

router = APIRouter()

# ── 需求追踪并发控制：(record_id, username) 粒度，支持多用户并发 ─────────────
import asyncio as _trac_asyncio
_trac_running: dict[tuple, str] = {}   # (record_id, username) → 'extract' | 'map'
_trac_lock = _trac_asyncio.Lock()


# ── 公共响应构建 ──────────────────────────────────────────────────────────────

import re as _re_step
_step_prefix = _re_step.compile(r'^\s*\d+\.\s*')


def _clean_steps(modules):
    cleaned = []
    for mod in (modules or []):
        mod = dict(mod)
        cases = []
        for case in mod.get("cases", []):
            case = dict(case)
            steps = case.get("steps", [])
            if isinstance(steps, list):
                case["steps"] = [_step_prefix.sub('', str(s)) for s in steps]
            cases.append(case)
        mod["cases"] = cases
        cleaned.append(mod)
    return cleaned


def _ai_case_response(record) -> AICaseFileResponse:
    modules = _clean_steps((record.cases_data or {}).get("modules", []))
    return AICaseFileResponse(
        id=record.id,
        task_name=record.task_name,
        case_count=record.case_count,
        has_md=bool(record.md_path),
        has_xmind=bool(record.xmind_path),
        modules=modules,
        created_at=record.created_at.isoformat() if record.created_at else "",
        doc_hash=record.doc_hash,
        parent_id=record.parent_id,
        diff_summary=record.diff_summary,
        record_status=getattr(record, "record_status", "active") or "active",
        gen_status=getattr(record, "gen_status", "done") or "done",
        gen_progress=getattr(record, "gen_progress", 0) or 0,
    )


# ── 后台生成任务 ──────────────────────────────────────────────────────────────

async def _do_generate_bg(
    record_id: int,
    task_name: str,
    document_path: Optional[str],
    content: Optional[str],
    formats: List[str],
) -> None:
    """AI 用例后台生成任务：生成完成后写库并通过 WebSocket 推送结果。"""
    from tools.database import async_session_maker
    from skills.ai_case_generator import ai_case_generator, release_generate_slot

    async with async_session_maker() as bg_db:
        res = await bg_db.execute(select(AICaseFile).where(AICaseFile.id == record_id))
        record = res.scalar_one_or_none()
        if not record:
            logger.error(f"_do_generate_bg: record_id={record_id} 不存在，任务中止")
            await release_generate_slot()
            return

        async def _progress(pct: int, stage: str):
            # 同步写库，前端重连后可从 GET /ai-cases/{id} 拿到当前进度
            try:
                record.gen_progress = pct
                await bg_db.commit()
            except Exception:
                pass  # 进度写库失败不阻断主流程
            await ws_manager.broadcast(
                {"type": "ai_gen_progress", "percent": pct, "stage": stage},
                client_id="ai_gen",
            )

        try:
            result = await ai_case_generator.generate(
                task_name=task_name,
                document_path=document_path,
                content=content,
                formats=formats,
                progress_cb=_progress,
                rag_source_id=record_id,
            )
            record.case_count  = result.get("case_count", 0)
            record.md_path     = result["files"].get("md")
            record.xmind_path  = result["files"].get("xmind")
            record.cases_data  = result.get("cases_data")
            record.doc_hash    = result.get("doc_hash")
            record.doc_content = result.get("doc_content")
            record.gen_status  = "done"
            await bg_db.commit()
            await bg_db.refresh(record)
            await ws_manager.broadcast(
                {"type": "ai_gen_done", "record_id": record_id,
                 "case_count": record.case_count, "task_name": task_name},
                client_id="ai_gen",
            )
            logger.info(f"AI 用例后台生成完成: record_id={record_id}，{record.case_count} 条")
        except Exception as e:
            record.gen_status = "failed"
            await bg_db.commit()
            logger.exception("AI 用例后台生成失败: record_id={}, err={}", record_id, repr(e))
            await ws_manager.broadcast(
                {"type": "ai_gen_progress", "percent": 0,
                 "stage": f"生成失败: {type(e).__name__}: {e}", "error": True},
                client_id="ai_gen",
            )
        finally:
            await release_generate_slot()


# ── 生成 / 列表 / 详情 / 下载 ─────────────────────────────────────────────────

@router.post("/ai-cases/generate", response_model=AICaseFileResponse)
@limiter.limit("5/minute")
async def generate_ai_cases(
    request: Request,
    body: AICaseGenerateRequest,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from skills.ai_case_generator import acquire_generate_slot, get_active_generate_count, _MAX_ACTIVE_GENERATE
    if not body.document_path and not body.content:
        raise HTTPException(status_code=400, detail="请提供文档路径或需求文本内容")
    # 校验工作空间成员身份
    if body.workspace_id:
        await check_workspace_member(db, body.workspace_id, current_user, "生成AI用例")
    # 并发任务上限检查
    if not await acquire_generate_slot():
        raise HTTPException(
            status_code=429,
            detail=f"当前已有 {get_active_generate_count()} 个生成任务进行中（上限 {_MAX_ACTIVE_GENERATE}），请稍后再试"
        )
    placeholder = AICaseFile(
        task_name=body.task_name, case_count=0, cases_data=None, gen_status="generating",
        created_by=current_user.username,
        project_id=body.workspace_id,
    )
    db.add(placeholder)
    await db.commit()
    await db.refresh(placeholder)

    # 优先走 ARQ（Redis 已配置且连接正常），否则降级为 BackgroundTasks
    from worker.arq_worker import get_arq_pool, task_generate_ai_cases
    arq_pool = get_arq_pool()
    if arq_pool:
        await arq_pool.enqueue_job(
            task_generate_ai_cases.__name__,
            record_id=placeholder.id, task_name=body.task_name,
            document_path=body.document_path, content=body.content, formats=body.formats,
        )
        logger.info(f"[ARQ] AI 用例生成任务已入队: record_id={placeholder.id}")
    else:
        background_tasks.add_task(
            _do_generate_bg,
            record_id=placeholder.id, task_name=body.task_name,
            document_path=body.document_path, content=body.content, formats=body.formats,
        )

    return _ai_case_response(placeholder)


@router.get("/ai-cases", response_model=List[AICaseFileResponse])
async def list_ai_cases(workspace_id: int = None, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    stmt = select(AICaseFile).order_by(AICaseFile.created_at.desc())
    stmt = stmt.where(AICaseFile.record_status == "active")
    f = await workspace_filter_members(db, AICaseFile, workspace_id, current_user)
    if f is not None:
        stmt = stmt.where(f)
    result = await db.execute(stmt)
    return [_ai_case_response(r) for r in result.scalars().all()]


@router.get("/ai-cases/{record_id}", response_model=AICaseFileResponse)
async def get_ai_case(record_id: int, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(AICaseFile).where(AICaseFile.id == record_id))
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    await check_access(db, record, current_user, "AI用例")
    return _ai_case_response(record)


@router.get("/ai-cases/{record_id}/download")
async def download_ai_case(record_id: int, format: str = "md", db: AsyncSession = Depends(get_db),
                          current_user: User = Depends(get_current_user)):
    from fastapi.responses import FileResponse
    from urllib.parse import quote
    from datetime import datetime
    result = await db.execute(select(AICaseFile).where(AICaseFile.id == record_id))
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    await check_access(db, record, current_user, "AI用例")
    if format == "md":
        file_path, media_type, ext = record.md_path, "text/markdown", ".md"
    elif format == "xmind":
        file_path, media_type, ext = record.xmind_path, "application/octet-stream", ".xmind"
    else:
        raise HTTPException(status_code=400, detail="不支持的格式，请使用 md 或 xmind")
    if not file_path:
        raise HTTPException(status_code=404, detail=f"该记录未生成 {format} 文件")

    # 解析路径
    p = Path(file_path)
    if not p.is_absolute():
        p = Path(__file__).parent.parent.parent / file_path

    # 文件不存在时，从 cases_data 重新生成
    if not p.exists():
        cases_data = record.cases_data
        if not cases_data or not cases_data.get("modules"):
            raise HTTPException(status_code=404, detail="文件已丢失且无法重新生成（用例数据为空）")
        try:
            from skills.ai_case_generator import ai_case_generator
            ts = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
            if format == "md":
                p = await ai_case_generator._save_markdown(cases_data, record.task_name, ts)
                record.md_path = str(p)
            else:
                p = await ai_case_generator._save_xmind(cases_data, record.task_name, ts)
                record.xmind_path = str(p)
            from sqlalchemy.orm.attributes import flag_modified
            flag_modified(record, "md_path" if format == "md" else "xmind_path")
            await db.commit()
            logger.info(f"下载时文件不存在，已重新生成: record_id={record_id}, format={format}, path={p}")
        except Exception as e:
            logger.error(f"重新生成下载文件失败: record_id={record_id}, err={e}")
            raise HTTPException(status_code=500, detail="文件已丢失，重新生成失败，请联系管理员")

    encoded = quote(record.task_name.replace("/", "_") + ext, safe="")
    return FileResponse(path=str(p), media_type=media_type,
                        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded}"})


@router.get("/ai-cases/{record_id}/export-excel")
async def export_ai_case_excel(
    record_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """将 AI 用例导出为 Excel，含完整用例字段 + 测试结果下拉列 + 备注列。"""
    from fastapi.responses import Response
    from urllib.parse import quote
    import io, openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    result = await db.execute(select(AICaseFile).where(AICaseFile.id == record_id))
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    await check_access(db, record, current_user, "AI用例")

    cases_data = record.cases_data or {}
    modules = cases_data.get("modules", [])

    # ── 收集所有有效用例（排除废弃）────────────────────────────────────────────
    rows = []
    for mod in modules:
        mod_name = mod.get("name", "通用")
        for case in mod.get("cases", []):
            if case.get("status") == "deprecated":
                continue
            steps = case.get("steps", [])
            steps_text = "\n".join(
                f"{i+1}. {s}" for i, s in enumerate(steps)
            ) if isinstance(steps, list) else str(steps)
            rows.append({
                "用例编号":   case.get("id", ""),
                "所属模块":   mod_name,
                "用例名称":   case.get("name", ""),
                "优先级":     case.get("priority", "P1"),
                "用例类型":   case.get("type", "功能测试"),
                "测试方法":   case.get("test_method", ""),
                "前置条件":   case.get("preconditions", ""),
                "测试步骤":   steps_text,
                "预期结果":   case.get("expected", ""),
                "测试结果":   "",   # 下拉列：Pass / Fail / NT / Block / NA
                "备注":       "",
            })

    # ── 创建工作簿 ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试用例"

    HEADERS = ["用例编号", "所属模块", "用例名称", "优先级", "用例类型",
               "测试方法", "前置条件", "测试步骤", "预期结果", "测试结果", "备注"]

    # ── 颜色和样式定义（颜色统一用 FFRRGGBB，避免 openpyxl 补 00 导致透明）────
    from openpyxl.styles import Color as _Color
    HEADER_BG   = "FF1D4ED8"   # 深蓝
    HEADER_FONT = Font(name="微软雅黑", bold=True, color=_Color(rgb="FFFFFFFF"), size=11)
    CELL_FONT   = Font(name="微软雅黑", size=10)
    ALT_BG      = PatternFill("solid", fgColor=_Color(rgb="FFEEF2FF"))   # 隔行浅蓝
    RESULT_BG   = PatternFill("solid", fgColor=_Color(rgb="FFFFF9E6"))   # 测试结果列淡黄
    NOTE_BG     = PatternFill("solid", fgColor=_Color(rgb="FFF0FDF4"))   # 备注列淡绿

    # 优先级色标（ARGB）
    PRIORITY_COLOR = {
        "P0": ("FFFF4D4F", "FFFFFFFF"),  # 红底白字
        "P1": ("FFFA8C16", "FFFFFFFF"),  # 橙底白字
        "P2": ("FF52C41A", "FFFFFFFF"),  # 绿底白字
    }

    thin = Side(style="thin", color="FFC7D2FE")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    # ── 写表头 ──────────────────────────────────────────────────────────────────
    ws.append(HEADERS)
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor=_Color(rgb=HEADER_BG))
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[1].height = 28

    # ── 写数据行（从第2行开始）───────────────────────────────────────────────
    for row_idx, data in enumerate(rows, 2):
        is_alt = (row_idx % 2 == 0)
        for col_idx, key in enumerate(HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = data[key]
            cell.font = CELL_FONT
            cell.border = border

            # 优先级列特殊着色
            if key == "优先级":
                pval = data[key]
                if pval in PRIORITY_COLOR:
                    bg, fg = PRIORITY_COLOR[pval]
                    cell.fill = PatternFill("solid", fgColor=_Color(rgb=bg))
                    cell.font = Font(name="微软雅黑", bold=True, color=_Color(rgb=fg), size=10)
                cell.alignment = center
            elif key == "测试结果":
                cell.fill = RESULT_BG
                cell.alignment = center
            elif key == "备注":
                cell.fill = NOTE_BG
                cell.alignment = wrap
            elif key in ("用例编号", "所属模块", "用例类型", "测试方法"):
                cell.alignment = center
                if is_alt:
                    cell.fill = ALT_BG
            else:
                cell.alignment = wrap
                if is_alt:
                    cell.fill = ALT_BG

        # 自动行高：按步骤行数估算
        step_lines = data["测试步骤"].count("\n") + 1
        ws.row_dimensions[row_idx].height = max(20, min(step_lines * 16, 120))

    # ── 测试结果列下拉验证（Pass / Fail / NT / NA / Block 五项）─────────────────
    result_col_letter = get_column_letter(HEADERS.index("测试结果") + 1)
    data_end_row = max(len(rows) + 1, 2)
    result_range = f"{result_col_letter}2:{result_col_letter}{data_end_row}"
    dv = DataValidation(
        type="list",
        formula1='"Pass,Fail,NT,NA,Block"',
        allow_blank=True,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="输入无效",
        error="请从下拉列表中选择：Pass / Fail / NT / NA / Block",
    )
    dv.sqref = result_range
    ws.add_data_validation(dv)

    # ── 测试结果单元格按值条件格式着色 ────────────────────────────────────────
    # 颜色统一用 FFRRGGBB，深背景+白字，醒目易读
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill as _PF, Font as _Ft, Color as _Color

    RESULT_STYLES = {
        "Pass":  {"bg": "FFFFFFFF", "fg": "FF389E0D"},   # 白底绿字
        "Fail":  {"bg": "FFFFFFFF", "fg": "FFF5222D"},   # 白底红字
        "NT":    {"bg": "FFFFFFFF", "fg": "FFD48806"},   # 白底黄字
        "NA":    {"bg": "FFFFFFFF", "fg": "FF8C8C8C"},   # 白底灰字
        "Block": {"bg": "FFFFFFFF", "fg": "FFD46B08"},   # 白底橙字
    }
    for val, style in RESULT_STYLES.items():
        fill = _PF("solid", fgColor=_Color(rgb=style["bg"]))
        font = _Ft(name="微软雅黑", bold=True, color=_Color(rgb=style["fg"]), size=13)
        rule = CellIsRule(
            operator="equal",
            formula=[f'"{val}"'],
            fill=fill,
            font=font,
        )
        ws.conditional_formatting.add(result_range, rule)

    # ── 自动筛选（AutoFilter）覆盖全部列 ──────────────────────────────────────
    last_col_letter = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A1:{last_col_letter}1"

    # ── 冻结首行 ────────────────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── 列宽 ────────────────────────────────────────────────────────────────────
    COL_WIDTHS = {
        "用例编号": 12, "所属模块": 14, "用例名称": 30, "优先级": 8,
        "用例类型": 12, "测试方法": 14, "前置条件": 22, "测试步骤": 45,
        "预期结果": 35, "测试结果": 12, "备注": 20,
    }
    for col_idx, header in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(header, 14)

    # ── 说明 Sheet：测试结果颜色图例（静态着色，让用户一眼看清颜色对应关系）────
    ws_legend = wb.create_sheet(title="测试结果说明")
    legend_items = [
        ("Pass",  "FFFFFFFF", "FF389E0D", "通过"),
        ("Fail",  "FFFFFFFF", "FFF5222D", "失败"),
        ("NT",    "FFFFFFFF", "FFD48806", "未测试"),
        ("NA",    "FFFFFFFF", "FF8C8C8C", "不适用"),
        ("Block", "FFFFFFFF", "FFD46B08", "阻塞"),
    ]
    # 标题行
    ws_legend.merge_cells("A1:C1")
    title_cell = ws_legend["A1"]
    title_cell.value = "测试结果颜色说明"
    title_cell.font = Font(name="微软雅黑", bold=True, size=13, color=_Color(rgb="FFFFFFFF"))
    title_cell.fill = PatternFill("solid", fgColor=_Color(rgb="FF1D4ED8"))
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_legend.row_dimensions[1].height = 30

    # 列头
    for ci, h in enumerate(["结果值", "含义", "颜色预览"], 1):
        c = ws_legend.cell(row=2, column=ci)
        c.value = h
        c.font = Font(name="微软雅黑", bold=True, size=10, color=_Color(rgb="FF333333"))
        c.fill = PatternFill("solid", fgColor=_Color(rgb="FFE8EFFE"))
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    ws_legend.row_dimensions[2].height = 22

    for ri, (val, bg, fg, meaning) in enumerate(legend_items, 3):
        # 结果值（带背景色）
        c_val = ws_legend.cell(row=ri, column=1, value=val)
        c_val.font = Font(name="微软雅黑", bold=True, size=13, color=_Color(rgb=fg))
        c_val.fill = PatternFill("solid", fgColor=_Color(rgb=bg))
        c_val.alignment = Alignment(horizontal="center", vertical="center")
        c_val.border = border

        # 含义
        c_mean = ws_legend.cell(row=ri, column=2, value=meaning)
        c_mean.font = Font(name="微软雅黑", size=10)
        c_mean.alignment = Alignment(horizontal="center", vertical="center")
        c_mean.border = border

        # 颜色预览（整格填色）
        c_preview = ws_legend.cell(row=ri, column=3, value="")
        c_preview.fill = PatternFill("solid", fgColor=_Color(rgb=bg))
        c_preview.border = border

        ws_legend.row_dimensions[ri].height = 28

    ws_legend.column_dimensions["A"].width = 12
    ws_legend.column_dimensions["B"].width = 12
    ws_legend.column_dimensions["C"].width = 18

    # ── 输出字节流 ──────────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = quote(record.task_name.replace("/", "_") + ".xlsx", safe="")
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


# ── 覆盖度分析 ────────────────────────────────────────────────────────────────

@router.get("/ai-cases/{record_id}/coverage")
async def get_ai_case_coverage(record_id: int, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(AICaseFile).where(AICaseFile.id == record_id))
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    await check_access(db, record, current_user, "AI用例")
    modules_data = (record.cases_data or {}).get("modules", [])
    all_cases = []
    for mod in modules_data:
        if all(c.get("status") == "deprecated" for c in mod.get("cases", []) if mod.get("cases")):
            continue
        for case in mod.get("cases", []):
            if case.get("status") != "deprecated":
                all_cases.append({**case, "_module": mod.get("name", "通用")})
    total = len(all_cases)
    if total == 0:
        return {"score": 0, "total": 0, "suggestions": ["当前无测试用例"]}
    priority_count = {"P0": 0, "P1": 0, "P2": 0}
    for c in all_cases:
        p = c.get("priority", "P1")
        priority_count[p] = priority_count.get(p, 0) + 1
    module_map: Dict[str, Dict] = {}
    for c in all_cases:
        m = c.get("_module", "通用")
        if m not in module_map:
            module_map[m] = {"total": 0, "P0": 0, "P1": 0, "P2": 0}
        p = c.get("priority", "P1")
        module_map[m]["total"] += 1
        module_map[m][p] = module_map[m].get(p, 0) + 1
    STANDARD_METHODS = ["等价类划分", "边界值分析", "判定表", "场景法", "错误推测", "状态转换"]
    used_methods = {m for c in all_cases for m in STANDARD_METHODS if m in (c.get("test_method") or "")}
    method_coverage = [{"name": m, "covered": m in used_methods} for m in STANDARD_METHODS]
    method_rate = round(len(used_methods) / len(STANDARD_METHODS) * 100)
    TYPES = ["功能测试", "性能测试", "兼容性测试"]
    type_count = {t: 0 for t in TYPES}
    for c in all_cases:
        for key in TYPES:
            if key in (c.get("type") or ""):
                type_count[key] += 1
                break
    p0_ratio = priority_count["P0"] / total
    score = max(0, min(100, round(
        len(used_methods) / len(STANDARD_METHODS) * 40
        + min(p0_ratio * 200, 30)
        + min(len(module_map) * 4, 20)
        + sum(1 for v in type_count.values() if v > 0) / len(TYPES) * 10
    )))
    suggestions = []
    missing = [m["name"] for m in method_coverage if not m["covered"]]
    if missing:
        suggestions.append(f"缺少测试方法：{'、'.join(missing)}，建议补充覆盖")
    if priority_count["P0"] == 0:
        suggestions.append("缺少 P0 核心用例，建议补充关键业务流程的主路径用例")
    elif p0_ratio < 0.1:
        suggestions.append(f"P0 用例仅占 {round(p0_ratio*100)}%，建议提升至 15% 以上")
    if priority_count["P2"] == 0:
        suggestions.append("缺少 P2 边界/异常用例，建议运用边界值分析和错误推测法补充")
    if type_count["性能测试"] == 0:
        suggestions.append("缺少性能测试用例，建议补充响应时间、并发等场景")
    if type_count["兼容性测试"] == 0:
        suggestions.append("缺少兼容性测试用例，建议补充多浏览器/多分辨率场景")
    for m_name, m_data in module_map.items():
        if m_data["P0"] == 0:
            suggestions.append(f"模块「{m_name}」无 P0 用例，建议补充核心流程")
    if not suggestions:
        suggestions.append("测试覆盖良好！可进一步增加状态转换和边界组合场景")
    return {
        "score": score, "total": total,
        "priority_distribution": priority_count,
        "module_distribution": [{"name": k, **v} for k, v in module_map.items()],
        "method_coverage": method_coverage, "method_rate": method_rate,
        "type_distribution": type_count, "suggestions": suggestions,
    }


# ── 优化 ──────────────────────────────────────────────────────────────────────

@router.post("/ai-cases/{record_id}/optimize", response_model=AICaseFileResponse)
@limiter.limit("3/minute")
async def optimize_ai_cases(request: Request, record_id: int, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    from skills.ai_case_generator import ai_case_generator
    result = await db.execute(select(AICaseFile).where(AICaseFile.id == record_id))
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    await check_access(db, record, current_user, "AI用例")
    cases_data = record.cases_data or {}
    if not cases_data.get("modules"):
        raise HTTPException(status_code=400, detail="该记录没有可优化的用例数据")

    async def _progress(pct: int, stage: str):
        await ws_manager.broadcast(
            {"type": "ai_gen_progress", "percent": pct, "stage": stage}, client_id="ai_gen",
        )

    try:
        opt_result = await ai_case_generator.optimize(
            task_name=record.task_name, cases_data=cases_data,
            formats=[f for f in ["md", "xmind"] if (f == "md" and record.md_path) or (f == "xmind" and record.xmind_path)],
            progress_cb=_progress,
        )
    except RuntimeError as e:
        await ws_manager.broadcast(
            {"type": "ai_gen_progress", "percent": 0, "stage": f"优化失败: {e}", "error": True},
            client_id="ai_gen",
        )
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("AI 用例优化失败: {}", repr(e))
        raise HTTPException(status_code=500, detail="用例优化失败，请稍后重试")

    record.cases_data = opt_result["cases_data"]
    record.case_count = opt_result["case_count"]
    if opt_result["files"].get("md"):
        record.md_path = opt_result["files"]["md"]
    if opt_result["files"].get("xmind"):
        record.xmind_path = opt_result["files"]["xmind"]
    await db.commit()
    await db.refresh(record)
    return _ai_case_response(record)


# ── 单条用例 CRUD ─────────────────────────────────────────────────────────────

@router.post("/ai-cases/{record_id}/cases", response_model=AICaseFileResponse)
async def add_ai_case_item(record_id: int, data: dict, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(AICaseFile).where(AICaseFile.id == record_id))
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    await check_access(db, record, current_user, "AI用例")
    cases_data = dict(record.cases_data or {})
    modules = list(cases_data.get("modules", []))
    module_name = (data.get("module") or "通用").strip()
    target_mod = next((m for m in modules if m["name"] == module_name), None)
    if target_mod is None:
        target_mod = {"name": module_name, "cases": []}
        modules.append(target_mod)
    all_nums = []
    for mod in modules:
        for c in mod.get("cases", []):
            cid = c.get("id", "")
            if cid.upper().startswith("TC"):
                try:
                    all_nums.append(int(cid[2:]))
                except ValueError:
                    pass
    new_id = f"TC{max(all_nums, default=0) + 1:03d}"
    target_mod["cases"].append({
        "id": new_id, "name": data.get("name", "新用例"),
        "priority": data.get("priority", "P1"), "type": data.get("type", "功能测试"),
        "test_method": data.get("test_method", ""), "preconditions": data.get("preconditions", ""),
        "steps": data.get("steps", []), "expected": data.get("expected", ""),
    })
    cases_data["modules"] = modules
    record.cases_data = cases_data
    record.case_count = sum(len(m.get("cases", [])) for m in modules)
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(record, "cases_data")
    await db.commit()
    await db.refresh(record)
    return _ai_case_response(record)


@router.put("/ai-cases/{record_id}/cases/{case_id}", response_model=AICaseFileResponse)
async def update_ai_case_item(record_id: int, case_id: str, data: dict, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(AICaseFile).where(AICaseFile.id == record_id))
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    await check_access(db, record, current_user, "AI用例")
    cases_data = dict(record.cases_data or {})
    modules = list(cases_data.get("modules", []))
    found_mod_idx = found_case_idx = None
    for mi, mod in enumerate(modules):
        for ci, case in enumerate(mod.get("cases", [])):
            if case.get("id") == case_id:
                found_mod_idx, found_case_idx = mi, ci
                break
        if found_mod_idx is not None:
            break
    if found_mod_idx is None:
        raise HTTPException(status_code=404, detail=f"用例 {case_id} 不存在")
    updated_case = dict(modules[found_mod_idx]["cases"][found_case_idx])
    for field in ("name", "priority", "type", "test_method", "preconditions", "steps", "expected"):
        if field in data:
            updated_case[field] = data[field]
    new_module = (data.get("module") or "").strip()
    if new_module and new_module != modules[found_mod_idx]["name"]:
        modules[found_mod_idx]["cases"].pop(found_case_idx)
        target_mod = next((m for m in modules if m["name"] == new_module), None)
        if target_mod is None:
            target_mod = {"name": new_module, "cases": []}
            modules.append(target_mod)
        target_mod["cases"].append(updated_case)
        modules = [m for m in modules if m.get("cases")]
    else:
        modules[found_mod_idx]["cases"][found_case_idx] = updated_case
    cases_data["modules"] = modules
    record.cases_data = cases_data
    record.case_count = sum(len(m.get("cases", [])) for m in modules)
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(record, "cases_data")
    await db.commit()
    await db.refresh(record)
    return _ai_case_response(record)


@router.delete("/ai-cases/{record_id}/cases/{case_id}", response_model=AICaseFileResponse)
async def delete_ai_case_item(record_id: int, case_id: str, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    result = await db.execute(select(AICaseFile).where(AICaseFile.id == record_id))
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    await check_access(db, record, current_user, "AI用例")
    cases_data = dict(record.cases_data or {})
    modules = list(cases_data.get("modules", []))
    found = False
    for mod in modules:
        for i, case in enumerate(mod.get("cases", [])):
            if case.get("id") == case_id:
                mod["cases"].pop(i)
                found = True
                break
        if found:
            break
    if not found:
        raise HTTPException(status_code=404, detail=f"用例 {case_id} 不存在")
    modules = [m for m in modules if m.get("cases")]
    cases_data["modules"] = modules
    record.cases_data = cases_data
    record.case_count = sum(len(m.get("cases", [])) for m in modules)
    from sqlalchemy.orm.attributes import flag_modified
    flag_modified(record, "cases_data")
    await db.commit()
    await db.refresh(record)
    return _ai_case_response(record)


@router.delete("/ai-cases/{record_id}")
async def delete_ai_case(record_id: int, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    from sqlalchemy import delete as sql_delete
    result = await db.execute(select(AICaseFile).where(AICaseFile.id == record_id))
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    await check_access(db, record, current_user, "AI用例")

    # 收集整条版本链（当前记录 + 所有 deprecated 父版本）的文件路径一并清理
    ids_to_delete = [record_id]
    parent_id = record.parent_id
    while parent_id:
        pr = await db.execute(select(AICaseFile).where(AICaseFile.id == parent_id))
        parent = pr.scalar_one_or_none()
        if not parent:
            break
        ids_to_delete.append(parent.id)
        _delete_case_files(parent)
        parent_id = parent.parent_id

    _delete_case_files(record)
    await db.execute(sql_delete(AICaseFile).where(AICaseFile.id.in_(ids_to_delete)))
    await db.commit()
    logger.info(f"删除 AI 用例记录链: ids={ids_to_delete}")
    return {"message": "删除成功"}


def _delete_case_files(record) -> None:
    """删除一条 AICaseFile 记录对应的 md/xmind 文件（文件不存在时静默忽略）。"""
    for path_str in [record.md_path, record.xmind_path]:
        if not path_str:
            continue
        p = Path(path_str)
        if not p.is_absolute():
            p = Path(__file__).parent.parent.parent / path_str
        p.unlink(missing_ok=True)


# ── Diff 检测 ─────────────────────────────────────────────────────────────────

@router.post("/ai-cases/{record_id}/diff-check")
async def diff_check_ai_case(record_id: int, request: DiffCheckRequest, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    from skills.ai_case_generator import ai_case_generator
    result = await db.execute(select(AICaseFile).where(AICaseFile.id == record_id))
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    await check_access(db, record, current_user, "AI用例")
    if not request.new_content and not request.new_document_path:
        raise HTTPException(status_code=400, detail="请提供新版文档路径或文本内容")
    if request.new_document_path:
        try:
            from tools.document_parser import document_parser
            parsed = await document_parser.parse(request.new_document_path)
            new_content = parsed.get("content", "")
            try:
                from bs4 import BeautifulSoup
                import re as _re
                soup = BeautifulSoup(new_content, "html.parser")
                for tag in soup(["script", "style", "head", "meta", "link"]):
                    tag.decompose()
                cleaned = _re.sub(r'\n{3,}', '\n\n', soup.get_text(separator="\n", strip=True)).strip()
                if len(cleaned) >= 200:
                    new_content = cleaned
            except Exception as _e:
                logger.debug(f"diff-check HTML 清洗跳过: {_e}")
            if len(new_content) > 100000:
                new_content = new_content[:100000]
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"新文档解析失败: {e}")
    else:
        new_content = request.new_content or ""
    if not new_content.strip():
        raise HTTPException(status_code=400, detail="新文档内容为空")
    new_doc_hash = ai_case_generator._compute_doc_hash(new_content)
    old_doc_hash = record.doc_hash or ""
    if old_doc_hash and new_doc_hash == old_doc_hash:
        return {"has_change": False, "new_doc_hash": new_doc_hash, "old_doc_hash": old_doc_hash,
                "diff": None, "message": "文档内容未发生变化，无需更新用例"}
    old_content = record.doc_content or ""
    if not old_content:
        return {"has_change": True, "new_doc_hash": new_doc_hash, "old_doc_hash": old_doc_hash,
                "diff": None, "message": "旧版文档内容未保存，无法做精确 Diff，可直接重新生成"}
    existing_module_names = [
        m.get("name", "") for m in (record.cases_data or {}).get("modules", [])
        if m.get("name") and "废弃" not in m.get("name", "")
    ]
    try:
        diff_result = await ai_case_generator.analyze_document_diff(
            old_doc_content=old_content, new_doc_content=new_content,
            existing_module_names=existing_module_names,
        )
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    return {"has_change": True, "new_doc_hash": new_doc_hash, "old_doc_hash": old_doc_hash, "diff": diff_result}


# ── 增量更新 ──────────────────────────────────────────────────────────────────

@router.post("/ai-cases/{record_id}/incremental-update", response_model=AICaseFileResponse)
async def incremental_update_ai_case(
    record_id: int, request: IncrementalUpdateRequest, db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from skills.ai_case_generator import ai_case_generator
    result = await db.execute(select(AICaseFile).where(AICaseFile.id == record_id))
    old_record = result.scalar_one_or_none()
    if not old_record:
        raise HTTPException(status_code=404, detail="记录不存在")
    await check_access(db, old_record, current_user, "AI用例")
    if not request.new_content and not request.new_document_path:
        raise HTTPException(status_code=400, detail="请提供新版文档路径或文本内容")
    if request.new_document_path:
        try:
            from tools.document_parser import document_parser
            parsed = await document_parser.parse(request.new_document_path)
            new_content = parsed.get("content", "")
            try:
                from bs4 import BeautifulSoup
                import re as _re
                soup = BeautifulSoup(new_content, "html.parser")
                for tag in soup(["script", "style", "head", "meta", "link"]):
                    tag.decompose()
                cleaned = _re.sub(r'\n{3,}', '\n\n', soup.get_text(separator="\n", strip=True)).strip()
                if len(cleaned) >= 200:
                    new_content = cleaned
            except Exception as _e:
                logger.debug(f"incremental-update HTML 清洗跳过: {_e}")
            if len(new_content) > 100000:
                new_content = new_content[:100000]
            p = Path(request.new_document_path)
            if "uploads" in p.parts and "documents" in p.parts:
                p.unlink(missing_ok=True)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"新文档解析失败: {e}")
    else:
        new_content = request.new_content or ""
    if not new_content.strip():
        raise HTTPException(status_code=400, detail="新文档内容为空")

    diff_result = request.diff
    existing_module_names = [
        m.get("name", "") for m in (old_record.cases_data or {}).get("modules", [])
        if m.get("name") and "废弃" not in m.get("name", "")
    ]
    if not diff_result:
        old_content = old_record.doc_content or ""
        if not old_content:
            raise HTTPException(status_code=400, detail="旧版文档内容未保存，无法做精确 Diff。请直接重新生成。")
        try:
            diff_result = await ai_case_generator.analyze_document_diff(
                old_doc_content=old_content, new_doc_content=new_content,
                existing_module_names=existing_module_names,
            )
        except RuntimeError as e:
            raise HTTPException(status_code=500, detail=str(e))

    if not (diff_result.get("changed") or diff_result.get("added") or diff_result.get("removed")):
        raise HTTPException(status_code=400, detail="Diff 分析未发现任何模块变更，无需更新")

    async def _progress(pct: int, stage: str):
        await ws_manager.broadcast(
            {"type": "ai_gen_progress", "percent": pct, "stage": stage}, client_id="ai_gen",
        )

    formats = [f for f in ["md", "xmind"]
               if (f == "md" and old_record.md_path) or (f == "xmind" and old_record.xmind_path)]
    if not formats:
        formats = ["md", "xmind"]

    try:
        upd_result = await ai_case_generator.incremental_update(
            task_name=old_record.task_name, old_cases_data=old_record.cases_data or {},
            new_doc_content=new_content, diff_result=diff_result,
            formats=formats, progress_cb=_progress,
        )
    except RuntimeError as e:
        await ws_manager.broadcast(
            {"type": "ai_gen_progress", "percent": 0, "stage": f"增量更新失败: {e}", "error": True},
            client_id="ai_gen",
        )
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("增量更新失败: {}", repr(e))
        raise HTTPException(status_code=500, detail="增量更新失败，请稍后重试")

    old_record.record_status = "deprecated"
    await db.flush()

    new_record = AICaseFile(
        task_name=old_record.task_name, case_count=upd_result.get("case_count", 0),
        md_path=upd_result["files"].get("md"), xmind_path=upd_result["files"].get("xmind"),
        cases_data=upd_result.get("cases_data"), doc_hash=upd_result.get("doc_hash"),
        doc_content=upd_result.get("doc_content"), parent_id=old_record.id,
        diff_summary=upd_result.get("diff_summary"), record_status="active",
        created_by=current_user.username,
        project_id=old_record.project_id,
    )
    db.add(new_record)
    await db.commit()
    await db.refresh(new_record)

    if new_content:
        import asyncio as _asyncio
        from skills.rag import index_document as _index_doc
        _asyncio.create_task(_index_doc(new_record.id, "ai_case", new_content))

    logger.info(f"增量更新完成: 旧记录 #{old_record.id} → 新记录 #{new_record.id}，{new_record.case_count} 条")
    return _ai_case_response(new_record)


# ── 需求追踪：提取需求条目（后台任务） ────────────────────────────────────────

async def _do_extract_requirements_bg(record_id: int, username: str) -> None:
    """后台提取需求条目，完成后通过 WebSocket 推送结果。"""
    from datetime import datetime as _dt
    from tools.database import async_session_maker
    from skills.ai_case_generator import ai_case_generator
    from skills.prompt_loader import get_system, render_user
    import json as _json

    ws_client = f"trac_gen_{record_id}_{username}"   # 每个用户独立推送频道

    async def _push(pct: int, stage: str, **kwargs):
        await ws_manager.broadcast(
            {"type": "trac_gen_progress", "task_type": "extract",
             "percent": pct, "stage": stage, **kwargs},
            client_id=ws_client,
        )

    async with _trac_lock:
        _trac_running[(record_id, username)] = 'extract'

    try:
        async with async_session_maker() as db:
            result = await db.execute(select(AICaseFile).where(AICaseFile.id == record_id))
            record = result.scalar_one_or_none()
            if not record:
                await _push(0, "记录不存在", error=True)
                return

            doc_content = record.doc_content or ""
            if not doc_content.strip():
                await _push(0, "该记录未保存需求文档内容，无法提取需求", error=True)
                return

            modules = (record.cases_data or {}).get("modules", [])
            modules_hint = "\n".join(f"  - {m.get('name','')}" for m in modules) or "  （无）"

            all_cases_flat = []
            for mod in modules:
                for case in mod.get("cases", []):
                    if case.get("status") != "deprecated":
                        all_cases_flat.append(case.get("name", ""))
            cases_hint = "\n".join(f"  - {n}" for n in all_cases_flat[:80]) or "  （无）"

            await _push(20, "正在调用 AI 提取需求条目，请稍候...")

            # 提取是单次 AI 调用，用心跳推进度条（20→88，每2s +2），避免进度卡死
            import asyncio as _asyncio
            _stop_heartbeat = _asyncio.Event()
            async def _heartbeat():
                pct = 22
                stages = [
                    "正在理解需求文档结构...",
                    "正在识别功能模块边界...",
                    "正在提取可测试的需求条目...",
                    "正在归类需求优先级...",
                    "正在生成需求 ID 和描述...",
                    "即将完成，请稍候...",
                ]
                idx = 0
                while not _stop_heartbeat.is_set() and pct < 88:
                    await _asyncio.sleep(2)
                    if _stop_heartbeat.is_set():
                        break
                    pct = min(pct + 3, 88)
                    stage = stages[min(idx, len(stages) - 1)]
                    idx += 1
                    await _push(pct, stage)
            heartbeat_task = _asyncio.create_task(_heartbeat())

            try:
                system_prompt = get_system("ai_case_gen.yaml", "extract_requirements")
                user_prompt = render_user("ai_case_gen.yaml", "extract_requirements",
                                          modules_hint=modules_hint,
                                          cases_hint=cases_hint,
                                          content=doc_content[:30000])
                raw = await ai_case_generator._run_claude_subprocess(system_prompt, user_prompt, timeout_secs=180)
                data = _json.loads(raw)
                requirements = data.get("requirements", [])
            except _json.JSONDecodeError as e:
                _stop_heartbeat.set()
                heartbeat_task.cancel()
                logger.error(f"需求提取返回非法 JSON: record_id={record_id}, err={e}")
                await _push(0, "AI 返回格式异常，请稍后重试", error=True)
                return
            except Exception as e:
                _stop_heartbeat.set()
                heartbeat_task.cancel()
                logger.exception("需求提取失败: record_id={}", record_id)
                await _push(0, f"需求提取失败: {e}", error=True)
                return
            finally:
                _stop_heartbeat.set()
                heartbeat_task.cancel()
                try:
                    await heartbeat_task
                except _asyncio.CancelledError:
                    pass

            if not requirements:
                await _push(0, "AI 未能从文档中提取到需求条目，请检查文档内容", error=True)
                return

            await _push(95, f"提取完成，共 {len(requirements)} 条需求，正在保存...")
            extracted_at = _dt.utcnow().isoformat() + "Z"
            record.requirements_data = {"extracted_at": extracted_at, "requirements": requirements}
            record.traceability_data = None
            from sqlalchemy.orm.attributes import flag_modified
            flag_modified(record, "requirements_data")
            flag_modified(record, "traceability_data")
            await db.commit()

            logger.info(f"需求提取完成: record_id={record_id}，共 {len(requirements)} 条需求")
            await ws_manager.broadcast(
                {"type": "trac_extract_done", "record_id": record_id,
                 "count": len(requirements), "extracted_at": extracted_at,
                 "requirements": requirements},
                client_id=ws_client,
            )

    finally:
        async with _trac_lock:
            if _trac_running.get((record_id, username)) == 'extract':
                _trac_running.pop((record_id, username), None)


@router.post("/ai-cases/{record_id}/extract-requirements")
@limiter.limit("3/minute")
async def extract_requirements(
    request: Request,
    record_id: int,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """从已保存的需求文档中提取结构化需求条目（后台任务，进度通过 WebSocket trac_gen 推送）。"""
    result = await db.execute(select(AICaseFile).where(AICaseFile.id == record_id))
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    await check_access(db, record, current_user, "AI用例")
    if not (record.doc_content or "").strip():
        raise HTTPException(
            status_code=400,
            detail="该记录未保存需求文档内容，无法提取需求（旧记录请重新生成用例后再使用此功能）"
        )
    async with _trac_lock:
        _trac_running[(record_id, current_user.username)] = 'extract'
    background_tasks.add_task(_do_extract_requirements_bg, record_id, current_user.username)
    return {"record_id": record_id, "status": "extracting", "message": "需求提取任务已启动，请通过 WebSocket 接收进度"}


# ── 需求追踪：用例-需求映射（后台任务） ──────────────────────────────────────

async def _do_map_cases_bg(record_id: int, username: str) -> None:
    """后台执行用例-需求映射，完成后通过 WebSocket 推送结果。"""
    from datetime import datetime as _dt
    from tools.database import async_session_maker
    from skills.ai_case_generator import ai_case_generator
    from skills.prompt_loader import get_system, render_user
    import json as _json

    ws_client = f"trac_gen_{record_id}_{username}"   # 每个用户独立推送频道

    async def _push(pct: int, stage: str, **kwargs):
        await ws_manager.broadcast(
            {"type": "trac_gen_progress", "task_type": "map",
             "percent": pct, "stage": stage, **kwargs},
            client_id=ws_client,
        )

    async with _trac_lock:
        _trac_running[(record_id, username)] = 'map'

    try:
        async with async_session_maker() as db:
            result = await db.execute(select(AICaseFile).where(AICaseFile.id == record_id))
            record = result.scalar_one_or_none()
            if not record:
                await _push(0, "记录不存在", error=True)
                return

            requirements = (record.requirements_data or {}).get("requirements", [])
            if not requirements:
                await _push(0, "请先提取需求条目", error=True)
                return

            modules = (record.cases_data or {}).get("modules", [])
            all_cases = []
            for mod in modules:
                for case in mod.get("cases", []):
                    if case.get("status") != "deprecated":
                        steps = case.get("steps", [])
                        steps_text = " → ".join(steps[:3]) if isinstance(steps, list) else str(steps)[:100]
                        expected = case.get("expected", case.get("expected_results", ""))
                        if isinstance(expected, list):
                            expected = "；".join(expected[:2])
                        all_cases.append({
                            "case_id": case.get("id", ""),
                            "name": case.get("name", ""),
                            "steps_summary": steps_text[:120],
                            "expected_summary": str(expected)[:80],
                        })

            if not all_cases:
                await _push(0, "该记录暂无有效用例", error=True)
                return

            requirements_text = "\n".join(
                f"  {r['id']} | {r['module']} | {r['title']} | {r.get('description','')} | 关键词：{','.join(r.get('keywords', []))}"
                for r in requirements
            )

            BATCH = 15
            total_batches = (len(all_cases) + BATCH - 1) // BATCH
            name_to_id = {c["name"]: c["case_id"] for c in all_cases}
            valid_ids = {c["case_id"] for c in all_cases}
            system_prompt = get_system("ai_case_gen.yaml", "map_cases_to_requirements")

            async def _map_batch(batch_num: int, batch: list, idx_to_id: dict):
                cases_text = "\n".join(
                    f"  {idx} | {c['name']} | {c['steps_summary']} | {c['expected_summary']}"
                    for idx, c in enumerate(batch)
                )
                user_prompt = render_user("ai_case_gen.yaml", "map_cases_to_requirements",
                                          requirements_text=requirements_text,
                                          cases_text=cases_text)
                try:
                    raw = await ai_case_generator._run_claude_subprocess(
                        system_prompt, user_prompt, timeout_secs=240)
                    data = _json.loads(raw)
                    batch_mappings = data.get("mappings", [])
                    fixed = []
                    for m in batch_mappings:
                        ai_id = str(m.get("case_id", ""))
                        real_id = (idx_to_id.get(ai_id)
                                   or (ai_id if ai_id in valid_ids else None)
                                   or name_to_id.get(ai_id))
                        if not real_id:
                            logger.debug(f"映射批次 {batch_num}: 无法匹配 case_id={ai_id!r}，跳过")
                            continue
                        raw_refs = m.get("req_refs", [])
                        req_refs = []
                        for ref in raw_refs:
                            if isinstance(ref, dict):
                                rid = ref.get("req_id", "")
                                conf = ref.get("confidence", "medium")
                                if conf != "low" and rid:
                                    req_refs.append(rid)
                            elif isinstance(ref, str) and ref:
                                req_refs.append(ref)
                        fixed.append({"case_id": real_id, "req_refs": req_refs})
                    non_empty = sum(1 for m in fixed if m.get("req_refs"))
                    logger.info(f"映射批次 {batch_num}: AI返回 {len(batch_mappings)} 条，"
                                f"有效 {len(fixed)} 条，有关联 {non_empty} 条")
                    return fixed
                except Exception as e:
                    logger.warning(f"映射批次 {batch_num} 失败（忽略）: {e}")
                    return []

            import asyncio as _asyncio
            batches = [
                (i // BATCH + 1,
                 all_cases[i:i + BATCH],
                 {str(idx): c["case_id"] for idx, c in enumerate(all_cases[i:i + BATCH])})
                for i in range(0, len(all_cases), BATCH)
            ]
            await _push(20, f"开始并行映射，共 {total_batches} 批（{len(all_cases)} 条用例）...")

            # 用 asyncio.as_completed 逐批收集结果，每完成一批实时推进度
            completed_batches = 0
            all_mappings: list = []
            tasks = {
                _asyncio.ensure_future(_map_batch(bn, b, idx)): bn
                for bn, b, idx in batches
            }
            for fut in _asyncio.as_completed(list(tasks.keys())):
                batch_result = await fut
                all_mappings.extend(batch_result)
                completed_batches += 1
                pct = 20 + int(completed_batches / total_batches * 72)
                has_refs = sum(1 for m in batch_result if m.get("req_refs"))
                await _push(
                    pct,
                    f"已完成 {completed_batches}/{total_batches} 批，"
                    f"本批 {len(batch_result)} 条用例，{has_refs} 条有关联需求"
                )
                await _push(
                    pct,
                    f"已完成 {completed_batches}/{total_batches} 批，"
                    f"本批 {len(batch_result)} 条用例，{has_refs} 条有关联需求"
                )

            if not all_mappings:
                await _push(0, "映射失败，请稍后重试", error=True)
                return

            await _push(95, f"全部 {total_batches} 批完成，正在保存映射结果...")
            mapped_at = _dt.utcnow().isoformat() + "Z"
            record.traceability_data = {"mapped_at": mapped_at, "mappings": all_mappings}
            from sqlalchemy.orm.attributes import flag_modified
            flag_modified(record, "traceability_data")
            await db.commit()

            logger.info(f"用例-需求映射完成: record_id={record_id}，{len(all_mappings)} 条用例已映射")
            await ws_manager.broadcast(
                {"type": "trac_map_done", "record_id": record_id,
                 "case_count": len(all_mappings), "req_count": len(requirements),
                 "mapped_at": mapped_at},
                client_id=ws_client,
            )

    finally:
        async with _trac_lock:
            if _trac_running.get((record_id, username)) == 'map':
                _trac_running.pop((record_id, username), None)


@router.post("/ai-cases/{record_id}/map-cases-to-reqs")
@limiter.limit("3/minute")
async def map_cases_to_requirements(
    request: Request,
    record_id: int,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """对现有用例做需求映射（后台任务，进度通过 WebSocket trac_gen 推送）。"""
    result = await db.execute(select(AICaseFile).where(AICaseFile.id == record_id))
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    await check_access(db, record, current_user, "AI用例")
    if not (record.requirements_data or {}).get("requirements"):
        raise HTTPException(status_code=400, detail="请先调用「提取需求」接口，生成需求列表后再进行映射")
    async with _trac_lock:
        _trac_running[(record_id, current_user.username)] = 'map'
    background_tasks.add_task(_do_map_cases_bg, record_id, current_user.username)
    return {"record_id": record_id, "status": "mapping", "message": "映射任务已启动，请通过 WebSocket 接收进度"}


# ── 需求追踪：追踪矩阵 ───────────────────────────────────────────────────────

@router.get("/ai-cases/{record_id}/traceability")
async def get_traceability(record_id: int, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    """返回完整需求-用例追踪矩阵。"""
    result = await db.execute(select(AICaseFile).where(AICaseFile.id == record_id))
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    await check_access(db, record, current_user, "AI用例")

    requirements_data = record.requirements_data or {}
    traceability_data = record.traceability_data or {}
    requirements = requirements_data.get("requirements", [])
    mappings = traceability_data.get("mappings", [])

    if not requirements:
        return {
            "ready":        False,
            "extracted_at": None,
            "mapped_at":    None,
            "requirements": [],
            "message":      "请先提取需求条目",
        }

    # 有需求但尚未映射：返回需求列表让前端正确恢复状态，ready=False
    if not traceability_data.get("mapped_at"):
        return {
            "ready":        False,
            "extracted_at": requirements_data.get("extracted_at"),
            "mapped_at":    None,
            "requirements": requirements,
            "matrix":       [],
            "orphan_cases": [],
            "summary":      {"total": len(requirements), "covered": 0,
                             "insufficient": 0, "uncovered": len(requirements), "coverage_rate": 0},
        }

    # case_id → req_refs 索引
    case_to_reqs: dict = {m["case_id"]: m.get("req_refs", []) for m in mappings}

    # req_id → 覆盖用例列表 索引
    modules = (record.cases_data or {}).get("modules", [])
    all_cases: dict = {}
    for mod in modules:
        for case in mod.get("cases", []):
            if case.get("status") != "deprecated":
                cid = case.get("id", "")
                all_cases[cid] = {
                    "id":       cid,
                    "name":     case.get("name", ""),
                    "module":   mod.get("name", ""),
                    "priority": case.get("priority", "P1"),
                    "req_refs": case_to_reqs.get(cid, []),
                }

    req_to_cases: dict = {r["id"]: [] for r in requirements}
    for cid, info in all_cases.items():
        for req_id in info["req_refs"]:
            if req_id in req_to_cases:
                req_to_cases[req_id].append({
                    "case_id":  cid,
                    "name":     info["name"],
                    "priority": info["priority"],
                })

    # 构建矩阵行
    matrix_rows = []
    for req in requirements:
        req_id = req["id"]
        linked = req_to_cases.get(req_id, [])
        count  = len(linked)
        status = "uncovered" if count == 0 else ("insufficient" if count == 1 else "covered")
        matrix_rows.append({
            "req_id":      req_id,
            "module":      req.get("module", ""),
            "title":       req.get("title", ""),
            "description": req.get("description", ""),
            "priority":    req.get("priority", "P1"),
            "case_count":  count,
            "status":      status,   # covered / insufficient / uncovered
            "cases":       linked,
        })

    total        = len(requirements)
    covered      = sum(1 for r in matrix_rows if r["status"] == "covered")
    insufficient = sum(1 for r in matrix_rows if r["status"] == "insufficient")
    uncovered    = sum(1 for r in matrix_rows if r["status"] == "uncovered")

    # 用例视角：未关联任何需求的用例
    orphan_cases = [
        {"case_id": cid, "name": info["name"], "module": info["module"]}
        for cid, info in all_cases.items()
        if not info["req_refs"]
    ]

    return {
        # ready 仅当映射真正执行过（mapped_at 有值）才为 True
        # 避免前端轮询在提取完成后误判"映射完成"
        "ready":        bool(traceability_data.get("mapped_at")),
        "extracted_at": requirements_data.get("extracted_at"),
        "mapped_at":    traceability_data.get("mapped_at"),
        "requirements": requirements,
        "summary": {
            "total":         total,
            "covered":       covered,
            "insufficient":  insufficient,
            "uncovered":     uncovered,
            "coverage_rate": round((covered + insufficient) / total * 100, 1) if total else 0,
        },
        "matrix":       matrix_rows,
        "orphan_cases": orphan_cases,
    }


# ── 需求追踪：分析覆盖缺口 ───────────────────────────────────────────────────

@router.post("/ai-cases/{record_id}/analyze-gap")
async def analyze_coverage_gap(record_id: int, data: dict, db: AsyncSession = Depends(get_db), current_user: User = Depends(get_current_user)):
    """
    分析指定需求的测试覆盖缺口。
    body: { req_id: str }
    返回：缺失维度列表 + 补充建议
    """
    from skills.ai_case_generator import ai_case_generator
    from skills.prompt_loader import get_system, render_user
    import json as _json

    result = await db.execute(select(AICaseFile).where(AICaseFile.id == record_id))
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    await check_access(db, record, current_user, "AI用例")

    req_id = data.get("req_id", "")
    if not req_id:
        raise HTTPException(status_code=400, detail="请提供 req_id")

    # 找到需求信息
    requirements = (record.requirements_data or {}).get("requirements", [])
    req = next((r for r in requirements if r["id"] == req_id), None)
    if not req:
        raise HTTPException(status_code=404, detail=f"需求 {req_id} 不存在")

    # 找到该需求已有的用例（从 traceability 反查）
    mappings = (record.traceability_data or {}).get("mappings", [])
    covered_case_ids = {
        m["case_id"] for m in mappings
        if req_id in m.get("req_refs", [])
    }
    # 从 cases_data 里取用例详情
    modules = (record.cases_data or {}).get("modules", [])
    existing_cases = []
    for mod in modules:
        for case in mod.get("cases", []):
            if case.get("id") in covered_case_ids and case.get("status") != "deprecated":
                existing_cases.append(case)

    existing_cases_text = "\n".join(
        f"  {c.get('id')} | {c.get('name')} | {c.get('test_method','')}"
        for c in existing_cases
    ) or "  （暂无关联用例）"

    system_prompt = get_system("ai_case_gen.yaml", "analyze_coverage_gap")
    user_prompt = render_user("ai_case_gen.yaml", "analyze_coverage_gap",
                              req_id=req_id,
                              req_title=req.get("title", ""),
                              req_description=req.get("description", ""),
                              req_priority=req.get("priority", "P1"),
                              existing_cases_text=existing_cases_text)
    try:
        raw = await ai_case_generator._run_claude_subprocess(system_prompt, user_prompt, timeout_secs=90)
        gap_data = _json.loads(raw)
    except _json.JSONDecodeError as e:
        logger.error(f"缺口分析返回非法 JSON: {e}")
        raise HTTPException(status_code=500, detail="AI 返回格式异常，请稍后重试")
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("缺口分析失败: record_id={}, req_id={}", record_id, req_id)
        raise HTTPException(status_code=500, detail="缺口分析失败，请稍后重试")

    return {
        "req_id":                req_id,
        "req_title":             req.get("title", ""),
        "existing_case_count":   len(existing_cases),
        "missing_dimensions":    gap_data.get("missing_dimensions", []),
        "supplement_suggestion": gap_data.get("supplement_suggestion", ""),
    }


# ── 需求追踪：生成补充用例（后台任务） ───────────────────────────────────────

async def _do_supplement_cases_bg(record_id: int, req_id: str, missing_dimensions: list, username: str = "") -> None:
    """后台生成补充用例，完成后通过 WebSocket 推送。"""
    from tools.database import async_session_maker
    from skills.ai_case_generator import ai_case_generator
    from skills.prompt_loader import get_system, render_user
    from sqlalchemy.orm.attributes import flag_modified
    import json as _json
    import re as _re
    import asyncio as _asyncio

    ws_client = f"trac_gen_{record_id}_{username}" if username else "trac_gen"

    async def _push(pct: int, stage: str, **kwargs):
        await ws_manager.broadcast(
            {"type": "trac_gen_progress", "percent": pct, "stage": stage, **kwargs},
            client_id=ws_client,
        )

    async with async_session_maker() as db:
        result = await db.execute(select(AICaseFile).where(AICaseFile.id == record_id))
        record = result.scalar_one_or_none()
        if not record:
            await _push(0, "记录不存在", error=True)
            return

        requirements = (record.requirements_data or {}).get("requirements", [])
        req = next((r for r in requirements if r["id"] == req_id), None)
        if not req:
            await _push(0, f"需求 {req_id} 不存在", error=True)
            return

        # 已有用例文本
        mappings = (record.traceability_data or {}).get("mappings", [])
        covered_ids = {m["case_id"] for m in mappings if req_id in m.get("req_refs", [])}
        modules = (record.cases_data or {}).get("modules", [])
        existing_cases = []
        for mod in modules:
            for case in mod.get("cases", []):
                if case.get("id") in covered_ids and case.get("status") != "deprecated":
                    existing_cases.append(case)

        existing_cases_text = "\n".join(
            f"  {c.get('id')} | {c.get('name')}"
            for c in existing_cases
        ) or "  （暂无）"

        missing_dimensions_text = "\n".join(
            f"  - {d['dimension']}：{d['reason']}\n    举例：{'; '.join(d.get('examples', []))}"
            for d in missing_dimensions
        )

        # 从 doc_content 中定位需求上下文
        doc_content = record.doc_content or ""
        doc_context = ""
        if doc_content:
            keywords = [req.get("title", ""), req_id, req.get("module", "")]
            for kw in keywords:
                if kw and len(kw) >= 2:
                    idx = doc_content.find(kw)
                    if idx != -1:
                        start = max(0, idx - 200)
                        end = min(len(doc_content), idx + 1000)
                        doc_context = doc_content[start:end]
                        break
            if not doc_context:
                doc_context = doc_content[:2000]

        await _push(30, f"正在为需求「{req.get('title', req_id)}」生成补充用例...")

        # AI 单次调用，加心跳避免进度卡死
        _stop_heartbeat = _asyncio.Event()
        async def _heartbeat():
            pct = 33
            stages = [
                "正在分析需求覆盖缺口...",
                "正在构建测试场景...",
                "正在生成测试步骤...",
                "正在完善预期结果...",
                "即将完成，请稍候...",
            ]
            idx = 0
            while not _stop_heartbeat.is_set() and pct < 88:
                await _asyncio.sleep(2)
                if _stop_heartbeat.is_set():
                    break
                pct = min(pct + 4, 88)
                await _push(pct, stages[min(idx, len(stages) - 1)])
                idx += 1
        heartbeat_task = _asyncio.create_task(_heartbeat())

        system_prompt = get_system("ai_case_gen.yaml", "generate_supplement_cases")
        user_prompt = render_user("ai_case_gen.yaml", "generate_supplement_cases",
                                  req_id=req_id,
                                  req_title=req.get("title", ""),
                                  req_description=req.get("description", ""),
                                  req_priority=req.get("priority", "P1"),
                                  existing_cases_text=existing_cases_text,
                                  missing_dimensions_text=missing_dimensions_text,
                                  doc_context=doc_context)
        try:
            raw = await ai_case_generator._run_claude_subprocess(system_prompt, user_prompt, timeout_secs=120)
            # 清理步骤前缀
            _step_prefix = _re.compile(r'^\s*\d+\.\s*')
            new_cases_raw = _json.loads(raw).get("cases", [])
            for case in new_cases_raw:
                if isinstance(case.get("steps"), list):
                    case["steps"] = [_step_prefix.sub('', str(s)) for s in case["steps"]]
        except Exception as e:
            logger.exception("补充用例生成失败: record_id={}, req_id={}", record_id, req_id)
            await _push(0, f"生成失败: {e}", error=True)
            return
        finally:
            _stop_heartbeat.set()
            heartbeat_task.cancel()
            try:
                await heartbeat_task
            except _asyncio.CancelledError:
                pass

        if not new_cases_raw:
            await _push(0, "AI 未能生成补充用例，请稍后重试", error=True)
            return

        # 找到目标模块（用需求的 module 字段）
        target_module_name = req.get("module", "通用")
        cases_data = dict(record.cases_data or {})
        mods = list(cases_data.get("modules", []))
        target_mod = next((m for m in mods if m["name"] == target_module_name), None)
        if not target_mod:
            target_mod = {"name": target_module_name, "cases": []}
            mods.append(target_mod)

        # 生成唯一 ID
        all_nums = []
        prefix_map = {
            "登录": "LOGIN", "注册": "REG", "用户": "USER", "权限": "AUTH",
            "支付": "PAY", "订单": "ORDER", "会员": "MEMBER", "首页": "HOME",
        }
        mod_prefix = next(
            (v for k, v in prefix_map.items() if k in target_module_name), None
        ) or "".join(c for c in target_module_name if c.strip())[:4].upper() or "SUP"

        for mod in mods:
            for c in mod.get("cases", []):
                cid = c.get("id", "")
                if cid.startswith(mod_prefix + "-"):
                    try:
                        all_nums.append(int(cid.split("-")[-1]))
                    except ValueError:
                        pass
        next_num = max(all_nums, default=0) + 1

        added_cases = []
        for case in new_cases_raw:
            case["id"] = f"{mod_prefix}-SUP-{next_num:03d}"
            case["is_supplement"] = True   # 标记为补充用例
            next_num += 1
            target_mod["cases"].append(case)
            added_cases.append({"id": case["id"], "name": case["name"]})

        cases_data["modules"] = mods
        record.cases_data = cases_data
        record.case_count = sum(len(m.get("cases", [])) for m in mods)
        flag_modified(record, "cases_data")

        # 更新 traceability_data：给新用例加映射
        trac = dict(record.traceability_data or {"mapped_at": "", "mappings": []})
        existing_mappings = list(trac.get("mappings", []))
        for case in new_cases_raw:
            existing_mappings.append({"case_id": case["id"], "req_refs": [req_id]})
        trac["mappings"] = existing_mappings
        record.traceability_data = trac
        flag_modified(record, "traceability_data")

        await db.commit()
        logger.info(f"补充用例生成完成: record_id={record_id}, req_id={req_id}, 新增 {len(added_cases)} 条")

        await _push(95, f"已生成 {len(added_cases)} 条补充用例，正在保存...")
        await ws_manager.broadcast(
            {"type": "trac_supplement_done", "record_id": record_id,
             "req_id": req_id, "added": added_cases, "count": len(added_cases)},
            client_id=ws_client,
        )


@router.post("/ai-cases/{record_id}/supplement-cases")
@limiter.limit("5/minute")
async def supplement_cases(
    request: Request,
    record_id: int,
    data: dict,
    background_tasks: BackgroundTasks,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    针对覆盖不足的需求生成补充用例（后台任务）。
    body: { req_id: str, missing_dimensions: [...] }
    """
    result = await db.execute(select(AICaseFile).where(AICaseFile.id == record_id))
    record = result.scalar_one_or_none()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    await check_access(db, record, current_user, "AI用例")

    req_id = data.get("req_id", "")
    missing_dimensions = data.get("missing_dimensions", [])
    if not req_id:
        raise HTTPException(status_code=400, detail="请提供 req_id")

    background_tasks.add_task(_do_supplement_cases_bg, record_id, req_id, missing_dimensions, current_user.username)
    return {"record_id": record_id, "req_id": req_id, "status": "generating",
            "message": "补充用例生成任务已启动，请通过 WebSocket 接收进度"}
